"""
ZenPlus Network Monitoring - Excel / CSV Export Service

Generates professional Excel (.xlsx) and CSV exports for network monitoring
reports, reusing the data-fetching layer from report_service.
"""

import csv
import io
from collections import Counter, defaultdict
from datetime import datetime, timezone
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.report_service import (
    _resolve_period,
    _fetch_devices,
    _fetch_service_checks,
    _fetch_alerts,
    _fetch_ping_metrics,
    _fetch_service_metrics,
    _fetch_device_status_log,
    _fetch_service_status_log,
    _device_uptime_pct,
    _device_rtt_stats,
    _service_uptime_pct,
    _mttr_seconds,
    _fmt_ms,
    _fmt_pct,
    _fmt_duration,
)

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_ALT_ROW_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
_DEFAULT_FONT = Font(name="Calibri", size=11)
_KPI_LABEL_FONT = Font(name="Calibri", bold=True, size=12)
_KPI_VALUE_FONT = Font(name="Calibri", bold=True, size=14, color="6366F1")
_TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="6366F1")
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="E2E8F0"),
)

VALID_REPORT_TYPES = {
    "executive_summary",
    "device_health",
    "service_health",
    "alert_analysis",
    "full_report",
}


# ---------------------------------------------------------------------------
# Worksheet helpers
# ---------------------------------------------------------------------------

def _apply_header_style(ws, num_cols: int) -> None:
    """Style the first row as a branded header and freeze it."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _apply_alternating_rows(ws) -> None:
    """Apply alternating white / light-gray background starting from row 2."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        if row_idx % 2 == 1:
            for cell in row:
                cell.fill = _ALT_ROW_FILL
        for cell in row:
            cell.font = _DEFAULT_FONT
            cell.border = _THIN_BORDER


def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Adjust column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        width = max(min_width, min(max_len + 3, max_width))
        ws.column_dimensions[col_letter].width = width


def _write_sheet(ws, headers: list[str], rows: list[list]) -> None:
    """Write headers + data rows, then apply styling."""
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _apply_header_style(ws, len(headers))
    _apply_alternating_rows(ws)
    _auto_fit_columns(ws)


def _write_kpi_sheet(ws, kpis: list[tuple[str, str]], title: str = "Report Summary") -> None:
    """Write a vertical KPI summary sheet with label/value pairs."""
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.cell(row=1, column=1).font = _TITLE_FONT
    ws.append([])  # spacer row

    for label, value in kpis:
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value=label).font = _KPI_LABEL_FONT
        ws.cell(row=row_num, column=2, value=value).font = _KPI_VALUE_FONT

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25


def _ts(val) -> str:
    """Format a datetime or string timestamp for display."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    return str(val)


# ---------------------------------------------------------------------------
# Data-assembly helpers (shared by Excel and CSV paths)
# ---------------------------------------------------------------------------

async def _collect_data(
    db: AsyncSession,
    report_type: str,
    period: str,
    from_time: Optional[datetime] = None,
    to_time: Optional[datetime] = None,
    device_ids: Optional[list[str]] = None,
    group_ids: Optional[list[str]] = None,
) -> dict:
    """Fetch all data needed for *report_type* and return as a dict."""

    if report_type not in VALID_REPORT_TYPES:
        raise ValueError(f"Unknown report type: {report_type!r}. "
                         f"Must be one of {VALID_REPORT_TYPES}")

    start, end, period_label = _resolve_period(period, from_time, to_time)

    data: dict = {
        "start": start,
        "end": end,
        "period_label": period_label,
    }

    need_devices = report_type in ("executive_summary", "device_health", "full_report")
    need_services = report_type in ("executive_summary", "service_health", "full_report")
    need_alerts = report_type in ("executive_summary", "alert_analysis", "full_report")
    need_device_detail = report_type in ("device_health", "full_report")
    need_service_detail = report_type in ("service_health", "full_report")

    # --- Devices ----------------------------------------------------------
    if need_devices or need_alerts:
        devices = await _fetch_devices(db, device_ids, group_ids)
        data["devices"] = devices
        d_ids = [d["id"] for d in devices]
        ping_rows = _fetch_ping_metrics(start, end, d_ids)
        data["ping_rows"] = ping_rows
    else:
        devices = []
        ping_rows = []

    # --- Services ---------------------------------------------------------
    if need_services or need_service_detail:
        services = await _fetch_service_checks(db)
        data["services"] = services
        svc_ids = [s["id"] for s in services]
        svc_rows = _fetch_service_metrics(start, end, svc_ids)
        data["svc_rows"] = svc_rows
    else:
        services = []
        svc_rows = []

    # --- Alerts -----------------------------------------------------------
    if need_alerts:
        d_ids = [d["id"] for d in devices]
        alerts = await _fetch_alerts(db, start, end, d_ids if d_ids else None)
        data["alerts"] = alerts

    # --- Status logs ------------------------------------------------------
    if need_device_detail:
        d_ids = [d["id"] for d in devices]
        data["device_status_log"] = _fetch_device_status_log(start, end, d_ids)

    if need_service_detail:
        svc_ids = [s["id"] for s in services]
        data["service_status_log"] = _fetch_service_status_log(start, end, svc_ids)

    return data


# ---------------------------------------------------------------------------
# Device / service row builders
# ---------------------------------------------------------------------------

def _build_device_rows(devices: list[dict], ping_rows: list[dict], *, full: bool = False) -> list[list]:
    """Build tabular rows for device overview."""
    rows = []
    for d in devices:
        uptime = _device_uptime_pct(ping_rows, d["id"])
        stats = _device_rtt_stats(ping_rows, d["id"])
        row = [
            d.get("hostname", ""),
            d.get("ip_address", ""),
            d.get("device_type", ""),
            d.get("group_name", ""),
        ]
        if full:
            row.append(d.get("location", ""))
        row += [
            d.get("status", ""),
            _fmt_pct(uptime),
            _fmt_ms(stats["avg"]),
        ]
        if full:
            row += [
                _fmt_ms(stats["min"]),
                _fmt_ms(stats["max"]),
                _fmt_ms(stats["p95"]),
            ]
        row.append(_ts(d.get("last_seen")))
        rows.append(row)
    return rows


def _device_headers(*, full: bool = False) -> list[str]:
    headers = ["Hostname", "IP Address", "Type", "Group"]
    if full:
        headers.append("Location")
    headers += ["Status", "Uptime %", "Avg RTT"]
    if full:
        headers += ["Min RTT", "Max RTT", "P95 RTT"]
    headers.append("Last Seen")
    return headers


def _build_service_rows(services: list[dict], svc_rows: list[dict]) -> list[list]:
    rows = []
    for s in services:
        uptime = _service_uptime_pct(svc_rows, s["id"])
        matching = [r for r in svc_rows if str(r.get("service_check_id")) == str(s["id"])]
        response_vals = [r["response_ms"] for r in matching if r.get("response_ms") is not None]
        avg_resp = sum(response_vals) / len(response_vals) if response_vals else None
        rows.append([
            s.get("name", ""),
            s.get("check_type", ""),
            s.get("target_url") or s.get("target_host", ""),
            s.get("status", ""),
            _fmt_pct(uptime),
            _fmt_ms(avg_resp),
            s.get("tls_days_remaining", ""),
        ])
    return rows


_SERVICE_HEADERS = ["Name", "Type", "Target", "Status", "Uptime %", "Avg Response", "TLS Days"]


def _build_alert_rows(alerts: list[dict]) -> list[list]:
    rows = []
    for a in alerts:
        resolved_at = a.get("resolved_at")
        triggered_at = a.get("triggered_at")
        if resolved_at and triggered_at:
            try:
                delta = (resolved_at - triggered_at).total_seconds()
                resolution = _fmt_duration(delta)
            except Exception:
                resolution = ""
        else:
            resolution = ""
        rows.append([
            _ts(a.get("triggered_at")),
            a.get("hostname", ""),
            a.get("severity", ""),
            a.get("status", ""),
            a.get("message", ""),
            _ts(a.get("acknowledged_at")),
            _ts(a.get("resolved_at")),
            resolution,
        ])
    return rows


_ALERT_HEADERS = [
    "Triggered At", "Device", "Severity", "Status", "Message",
    "Acknowledged At", "Resolved At", "Resolution Time",
]


# ---------------------------------------------------------------------------
# Excel sheet builders
# ---------------------------------------------------------------------------

def _add_executive_summary_sheets(wb: Workbook, data: dict) -> None:
    devices = data.get("devices", [])
    ping_rows = data.get("ping_rows", [])
    alerts = data.get("alerts", [])

    # --- Summary KPIs ---
    total = len(devices)
    online = sum(1 for d in devices if d.get("status") in ("online", "up"))
    online_pct = (online / total * 100) if total else 0
    rtt_vals = []
    for d in devices:
        stats = _device_rtt_stats(ping_rows, d["id"])
        if stats["avg"] is not None:
            rtt_vals.append(stats["avg"])
    avg_rtt = sum(rtt_vals) / len(rtt_vals) if rtt_vals else None
    mttr = _mttr_seconds(alerts)

    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_kpi_sheet(ws_summary, [
        ("Period", data["period_label"]),
        ("Total Devices", str(total)),
        ("Online %", _fmt_pct(online_pct)),
        ("Average RTT", _fmt_ms(avg_rtt)),
        ("Total Alerts", str(len(alerts))),
        ("MTTR", _fmt_duration(mttr)),
    ])

    # --- Device Status sheet ---
    ws_dev = wb.create_sheet("Device Status")
    _write_sheet(ws_dev, _device_headers(full=False),
                 _build_device_rows(devices, ping_rows, full=False))

    # --- Alert Summary sheet ---
    ws_alert = wb.create_sheet("Alert Summary")
    sev_counter = Counter(a.get("severity", "unknown") for a in alerts)
    sev_rows = [[sev, count] for sev, count in sev_counter.most_common()]
    _write_sheet(ws_alert, ["Severity", "Count"], sev_rows)

    # Top alerting devices (append below with a gap)
    device_counter = Counter(a.get("hostname", "unknown") for a in alerts)
    top_devices = device_counter.most_common(10)
    gap_row = ws_alert.max_row + 2
    ws_alert.cell(row=gap_row, column=1, value="Top Alerting Devices").font = _TITLE_FONT
    ws_alert.cell(row=gap_row + 1, column=1, value="Hostname").font = _HEADER_FONT
    ws_alert.cell(row=gap_row + 1, column=1).fill = _HEADER_FILL
    ws_alert.cell(row=gap_row + 1, column=2, value="Alert Count").font = _HEADER_FONT
    ws_alert.cell(row=gap_row + 1, column=2).fill = _HEADER_FILL
    for i, (hostname, count) in enumerate(top_devices):
        r = gap_row + 2 + i
        ws_alert.cell(row=r, column=1, value=hostname).font = _DEFAULT_FONT
        ws_alert.cell(row=r, column=2, value=count).font = _DEFAULT_FONT
        if i % 2 == 1:
            ws_alert.cell(row=r, column=1).fill = _ALT_ROW_FILL
            ws_alert.cell(row=r, column=2).fill = _ALT_ROW_FILL


def _add_device_health_sheets(wb: Workbook, data: dict, *, is_standalone: bool = True) -> None:
    devices = data.get("devices", [])
    ping_rows = data.get("ping_rows", [])

    # --- Device Overview ---
    if is_standalone:
        ws = wb.active
        ws.title = "Device Overview"
    else:
        ws = wb.create_sheet("Device Overview")
    _write_sheet(ws, _device_headers(full=True),
                 _build_device_rows(devices, ping_rows, full=True))

    # --- Group Summary ---
    groups: dict[str, list[dict]] = defaultdict(list)
    for d in devices:
        groups[d.get("group_name") or "Ungrouped"].append(d)
    group_rows = []
    for name, members in sorted(groups.items()):
        total = len(members)
        online = sum(1 for m in members if m.get("status") in ("online", "up"))
        pct = (online / total * 100) if total else 0
        group_rows.append([name, total, online, _fmt_pct(pct)])
    ws_grp = wb.create_sheet("Group Summary")
    _write_sheet(ws_grp, ["Group", "Total Devices", "Online", "Online %"], group_rows)

    # --- Status Changes ---
    status_log = data.get("device_status_log", [])
    hostname_map = {d["id"]: d.get("hostname", "") for d in devices}
    sc_rows = []
    for entry in status_log:
        sc_rows.append([
            _ts(entry.get("timestamp")),
            hostname_map.get(str(entry.get("device_id")), str(entry.get("device_id", ""))),
            entry.get("old_status", ""),
            entry.get("new_status", ""),
            entry.get("reason", ""),
            _fmt_duration(entry.get("duration_sec")),
        ])
    ws_sc = wb.create_sheet("Status Changes")
    _write_sheet(ws_sc, ["Timestamp", "Device", "Old Status", "New Status", "Reason", "Duration"],
                 sc_rows)


def _add_service_health_sheets(wb: Workbook, data: dict, *, is_standalone: bool = True) -> None:
    services = data.get("services", [])
    svc_rows_data = data.get("svc_rows", [])

    # --- Service Overview ---
    if is_standalone:
        ws = wb.active
        ws.title = "Service Overview"
    else:
        ws = wb.create_sheet("Service Overview")
    _write_sheet(ws, _SERVICE_HEADERS, _build_service_rows(services, svc_rows_data))

    # --- Service Metrics ---
    ws_met = wb.create_sheet("Service Metrics")
    svc_name_map = {str(s["id"]): s.get("name", "") for s in services}
    met_rows = []
    for r in svc_rows_data:
        met_rows.append([
            svc_name_map.get(str(r.get("service_check_id")), ""),
            _ts(r.get("timestamp")),
            _fmt_ms(r.get("response_ms")),
            "Yes" if r.get("is_up") else "No",
            r.get("status_code", ""),
            r.get("error_message", ""),
        ])
    _write_sheet(ws_met,
                 ["Service", "Timestamp", "Response", "Up", "Status Code", "Error"],
                 met_rows)

    # --- Service Status Changes ---
    status_log = data.get("service_status_log", [])
    sc_rows = []
    for entry in status_log:
        sc_rows.append([
            _ts(entry.get("timestamp")),
            svc_name_map.get(str(entry.get("service_check_id")), str(entry.get("service_check_id", ""))),
            entry.get("old_status", ""),
            entry.get("new_status", ""),
            entry.get("reason", ""),
            _fmt_duration(entry.get("duration_sec")),
        ])
    ws_sc = wb.create_sheet("Service Status Changes")
    _write_sheet(ws_sc, ["Timestamp", "Service", "Old Status", "New Status", "Reason", "Duration"],
                 sc_rows)


def _add_alert_analysis_sheets(wb: Workbook, data: dict, *, is_standalone: bool = True) -> None:
    alerts = data.get("alerts", [])

    # --- Alert List ---
    if is_standalone:
        ws = wb.active
        ws.title = "Alert List"
    else:
        ws = wb.create_sheet("Alert List")
    _write_sheet(ws, _ALERT_HEADERS, _build_alert_rows(alerts))

    # --- Severity Summary ---
    sev_counter = Counter(a.get("severity", "unknown") for a in alerts)
    total_alerts = len(alerts)
    sev_rows = []
    for sev, count in sev_counter.most_common():
        pct = (count / total_alerts * 100) if total_alerts else 0
        sev_rows.append([sev, count, _fmt_pct(pct)])
    ws_sev = wb.create_sheet("Severity Summary")
    _write_sheet(ws_sev, ["Severity", "Count", "Percentage"], sev_rows)

    # --- Top Alerting Devices ---
    device_counter = Counter(a.get("hostname", "unknown") for a in alerts)
    # Approximate total downtime per device from resolved alerts
    device_downtime: dict[str, float] = defaultdict(float)
    for a in alerts:
        hostname = a.get("hostname", "unknown")
        triggered = a.get("triggered_at")
        resolved = a.get("resolved_at")
        if triggered and resolved:
            try:
                device_downtime[hostname] += (resolved - triggered).total_seconds()
            except Exception:
                pass
    top_rows = []
    for hostname, count in device_counter.most_common(20):
        top_rows.append([
            hostname,
            count,
            _fmt_duration(device_downtime.get(hostname) or 0),
        ])
    ws_top = wb.create_sheet("Top Alerting Devices")
    _write_sheet(ws_top, ["Hostname", "Alert Count", "Total Downtime"], top_rows)


def _add_full_report_sheets(wb: Workbook, data: dict) -> None:
    """Combine all report type sheets into one workbook."""
    # The first sheet is a summary (executive)
    _add_executive_summary_sheets(wb, data)
    _add_device_health_sheets(wb, data, is_standalone=False)
    _add_service_health_sheets(wb, data, is_standalone=False)
    _add_alert_analysis_sheets(wb, data, is_standalone=False)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

async def generate_excel_report(
    db: AsyncSession,
    report_type: str,
    period: str,
    from_time: Optional[datetime] = None,
    to_time: Optional[datetime] = None,
    device_ids: Optional[list[str]] = None,
    group_ids: Optional[list[str]] = None,
) -> bytes:
    """Generate a styled Excel (.xlsx) workbook and return it as bytes."""

    data = await _collect_data(db, report_type, period, from_time, to_time,
                               device_ids, group_ids)

    wb = Workbook()

    if report_type == "executive_summary":
        _add_executive_summary_sheets(wb, data)
    elif report_type == "device_health":
        _add_device_health_sheets(wb, data, is_standalone=True)
    elif report_type == "service_health":
        _add_service_health_sheets(wb, data, is_standalone=True)
    elif report_type == "alert_analysis":
        _add_alert_analysis_sheets(wb, data, is_standalone=True)
    elif report_type == "full_report":
        _add_full_report_sheets(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def generate_csv_report(
    db: AsyncSession,
    report_type: str,
    period: str,
    from_time: Optional[datetime] = None,
    to_time: Optional[datetime] = None,
    device_ids: Optional[list[str]] = None,
    group_ids: Optional[list[str]] = None,
) -> bytes:
    """Generate a CSV export and return it as UTF-8 encoded bytes."""

    data = await _collect_data(db, report_type, period, from_time, to_time,
                               device_ids, group_ids)

    sio = io.StringIO()
    writer = csv.writer(sio)

    if report_type in ("executive_summary", "device_health", "full_report"):
        devices = data.get("devices", [])
        ping_rows = data.get("ping_rows", [])
        full = report_type in ("device_health", "full_report")
        writer.writerow(_device_headers(full=full))
        writer.writerows(_build_device_rows(devices, ping_rows, full=full))

    elif report_type == "service_health":
        services = data.get("services", [])
        svc_rows_data = data.get("svc_rows", [])
        writer.writerow(_SERVICE_HEADERS)
        writer.writerows(_build_service_rows(services, svc_rows_data))

    elif report_type == "alert_analysis":
        alerts = data.get("alerts", [])
        writer.writerow(_ALERT_HEADERS)
        writer.writerows(_build_alert_rows(alerts))

    return sio.getvalue().encode("utf-8")
